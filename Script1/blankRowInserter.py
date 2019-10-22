#! python3
#takes two integers and a filename string as command line arguments. 
# Let’s call the first integer N and the second integer M. 
# Starting at row N, the program should insert M blank rows into the spreadsheet

import openpyxl,sys
from openpyxl.utils import column_index_from_string,get_column_letter

file='Sample.xlsx'
n=2
m=3
wb=openpyxl.load_workbook(file)
sheet=wb.active
column=sheet.max_column+1
my_wb=openpyxl.Workbook()
my_sheet=my_wb.active
for i in range(1,sheet.max_row+1):
    if i >= n:
        temp=i+m
    else:
        temp=i
    for j in range(1,column):
        my_sheet[get_column_letter(j)+str(temp)]=sheet[get_column_letter(j)+str(i)].value        
print("Why stop here")
my_wb.save('nananana.xlsx')
