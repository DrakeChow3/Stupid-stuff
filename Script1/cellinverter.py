#! python3
# Write a program to invert the row and column of the cells in the spreadsheet

import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string

file="Sample.xlsx"
wb=openpyxl.load_workbook(file)
sheet=wb.active
column=sheet.max_column+1
my_wb=openpyxl.Workbook()
my_sheet=my_wb.active
for i in range(1,sheet.max_row+1):
    for j in range(1,column+1):
        my_sheet[get_column_letter(i)+str(j)]=sheet[get_column_letter(j)+str(i)].value        
print("Why stop here")
my_wb.save('nananana.xlsx')