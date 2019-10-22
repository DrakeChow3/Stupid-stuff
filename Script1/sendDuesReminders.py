#! python3
# sendDuesReminders.py -Sends emails based on payment status in spreadsheet.
#yup

import sys,openpyxl,smtplib,logging
logging.basicConfig(level=logging.DEBUG,format=' %(asctime)s - %(levelname)s - %(message)s ')
from openpyxl.utils import get_column_letter

openpyObj=openpyxl.load_workbook('duesRecords.xlsx')
sheet=openpyObj.active
maxCol=sheet.max_column
logging.debug('max column:'+str(maxCol))
lastMonth=sheet[get_column_letter(maxCol)+'1'].value
logging.debug('Latest month:'+lastMonth)
#create a dictionary to save who hasn't paid dues in the lastest month
c_dict={}
for r in range(2,sheet.max_row+1):
    state=sheet[get_column_letter(maxCol)+str(r)].value #get the lastest value in the lastest month of that person
    if state != 'paid':
        name=sheet[get_column_letter(1)+str(r)].value
        email=sheet[get_column_letter(2)+str(r)].value
        c_dict[name]=email
smtpObj=smtplib.SMTP('smtp.gmail.com','587')
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('chaunhatlong01061999@gmail.com',)
for n,v in c_dict.items():
    subject="""Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues for %s. 
    Please make this payment as soon as possible. Thank you!'""" %(lastMonth,n,lastMonth)
    print('Sending email to '+n)
    s_status=smtpObj.sendmail('chaunhatlong01061999@gmail.com',v,subject)
    if s_status != {}:
        print('There is an error sending email to '+n)
smtpObj.quit()