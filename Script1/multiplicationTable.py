#! python3
# Create a program that takes a number N 
# from the command line and creates an N×N multiplication table in an Excel spreadsheet

import openpyxl,sys
from openpyxl.utils import get_column_letter, column_index_from_string

n=int(sys.argv[1])
wb=openpyxl.Workbook()
sheet=wb.active
sheet.title="The NxN"
for i in range(1,n+2):
    for j in range(1,n+1):
        if i==1:
            sheet[get_column_letter(j+1)+str(i)]=j
        else:
            sheet['A'+str(i)]=i-1
            # print((sheet[get_column_letter(j+1)+'1'].value))
            # print(str(sheet[get_column_letter(j+1)+str(i)])+'='+str(sheet[get_column_letter(j+1)+'1'].value)+'*'+str(sheet['A'+str(i)].value))
            sheet[get_column_letter(j+1)+str(i)]=int(sheet[get_column_letter(j+1)+'1'].value)*int(sheet['A'+str(i)].value)
wb.save('NandN.xlsx')