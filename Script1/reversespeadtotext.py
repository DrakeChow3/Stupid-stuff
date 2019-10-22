#! python3
# reserve spreadsheet to text
# => text to spreadsheet

import openpyxl
from openpyxl.utils import get_column_letter

#in this example only keep 10 textfile or it would be too many to handle
wb=openpyxl.Workbook()
sheet=wb.active
for i in range(1,11):
    file=open('text'+str(i)+'.txt','r')
    file_list=file.readlines()
    for j in range(0,len(file_list)):
        sheet[get_column_letter(j+1)+str(i)]=file_list[j]
wb.save('hahaha.xlsx')