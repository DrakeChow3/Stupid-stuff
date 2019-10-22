#! python3
#write a program that reads all the Excel files in 
#the current working directory and outputs them as CSV files.

import os,openpyxl,csv,logging
from openpyxl.utils import get_column_letter
logging.basicConfig(level=logging.DEBUG,format=' %(asctime)s - %(levelname)s - %(message)s')

#make folder ExceltoCSV containg all the csv file
os.makedirs('ExceltoCSV',exist_ok=True)
#loop through all the file in the current directory
for file in os.listdir('.'):
    #find file with extension xlsx(Excel file)
    if file.endswith('.xlsx'):
        #load the excel file
        wb=openpyxl.load_workbook(file)
        #find all the sheet in the excel file
        for sheet_number in range(len(wb.sheetnames)):
            #find the current sheet name and get the sheet object
            sheet_title=wb.sheetnames[sheet_number]
            logging.debug('Sheet title is '+sheet_title)
            sheet=wb[sheet_title]
            #Creating the csv file associating with sheet's name
            out_name=open(os.path.join('ExceltoCSV',file[:-5]+'_'+sheet_title+'.csv'),'w',newline='')
            outputFile=csv.writer(out_name)
            #load all the row to a list and add the list to writerow
            for i in range(1,sheet.max_row+1):
                my_list=[]
                for j in range(1,sheet.max_column+1):
                    my_list.append(sheet[get_column_letter(j)+str(i)].value)
                logging.debug(my_list)
                outputFile.writerow(my_list)
            out_name.close()            