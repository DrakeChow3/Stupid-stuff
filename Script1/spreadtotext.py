#! python3
# spreadsheet to textfile

import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string

wb=openpyxl.load_workbook('Sample.xlsx')
sheet=wb.active
for i in range(1,sheet.max_row+1):
    my_list=[]
    for j in range(1,sheet.max_column+1):
        my_list.append(str(sheet[get_column_letter(j)+str(i)].value))
    text="\n".join(my_list)
    file=open('text'+str(i)+'.txt','w')
    file.write(text)
    file.close()